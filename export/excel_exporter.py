"""
Excel exporter with full openpyxl formatting.

Features
--------
- Sheet "Bets"    : colour-coded by tier (S/A/B/C) with bold headers,
                    freeze pane on row 1, auto-filter, auto-column widths
- Sheet "Summary" : league-level breakdown (bets, avg EV, avg Kelly, bet rate)
- Sheet "Metrics" : backtest KPIs table + sparklines text
- Conditional fill: S=green, A=cyan, B=yellow, C=light-grey, NO_BET=light-red
- Number formats  : probability (0.0%), odds (0.00), EV (+0.000%), kelly (0.000%)
- Date-stamped auto-naming
- Legacy export_value_bets() method for backwards compatibility

Usage
-----
    from export.excel_exporter import ExcelExporter

    exp = ExcelExporter(output_dir="outputs")
    path = exp.export(ranked_bets, metrics=backtest_metrics)
    path = exp.export_value_bets("outputs/bets.xlsx", bets)  # legacy
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Optional

_OPENPYXL_AVAILABLE = False
try:
    import openpyxl
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

_HEADER_FILL = "1A6BB5"  # deep blue
_HEADER_FONT = "FFFFFF"

_TIER_FILL = {
    "S": "C6EFCE",  # green
    "A": "DDEBF7",  # light blue
    "B": "FFEB9C",  # yellow
    "C": "F2F2F2",  # light grey
    "X": "FCE4D6",  # light red / no value
}
_BET_FILL = "C6EFCE"
_WATCHLIST_FILL = "FFEB9C"
_NO_BET_FILL = "FCE4D6"
_ALT_FILL = "F7FBFF"  # alternating row colour

# ---------------------------------------------------------------------------
# Field definitions
# ---------------------------------------------------------------------------

_BET_COLUMNS = [
    # (header, key, number_format, width)
    ("Match", "match_id", "@", 28),
    ("Market", "market", "@", 10),
    ("Probability", "probability", "0.00%", 14),
    ("Odds", "odds", "0.00", 8),
    ("EV", "ev", "+0.000%;[Red]-0.000%", 10),
    ("Kelly %", "kelly", "0.000%", 10),
    ("Stake €", "kelly_stake", "#,##0.00", 10),
    ("Tier", "tier", "@", 6),
    ("Confidence", "confidence", "0.0%", 12),
    ("Agreement", "agreement", "0.0%", 12),
    ("Market Edge", "market_edge", "+0.000%;[Red]-0.000%", 13),
    ("Decision", "decision", "@", 12),
    ("Over 2.5", "over_25", "0.0%", 10),
    ("BTTS Yes", "btts_yes", "0.0%", 10),
]

_SUMMARY_COLUMNS = [
    ("League", "league", "@", 24),
    ("Bets", "bets", "0", 8),
    ("Avg EV", "avg_ev", "+0.000%", 12),
    ("Avg Kelly", "avg_kelly", "0.000%", 12),
    ("Bet Rate", "bet_rate", "0.0%", 10),
]

_METRICS_COLUMNS = [
    ("KPI", "kpi", "@", 24),
    ("Value", "value", "@", 18),
]


# ---------------------------------------------------------------------------
# Helper: safe value conversion
# ---------------------------------------------------------------------------


def _num(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _int(v, default=0):
    try:
        return int(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _str(v) -> str:
    return str(v) if v is not None else ""


# ---------------------------------------------------------------------------
# Internal openpyxl helpers
# ---------------------------------------------------------------------------


def _make_fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _header_style(ws, row: int, ncols: int) -> None:
    """Apply bold white-on-blue style to a header row."""
    fill = _make_fill(_HEADER_FILL)
    font = Font(bold=True, color=_HEADER_FONT, size=10)
    thin = Side(border_style="thin", color="AAAAAA")
    bdr = Border(bottom=thin)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = bdr
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False
        )


def _row_fill(ws, row: int, ncols: int, hex_color: str) -> None:
    fill = _make_fill(hex_color)
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = fill


def _auto_width(ws, columns: list[tuple]) -> None:
    """Set column widths based on column definition hints."""
    for col_idx, (_, _, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_bets_sheet(ws, bets: list[dict]) -> None:
    """Fill the Bets worksheet."""
    # Header
    for col_idx, (header, _, fmt, _) in enumerate(_BET_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    _header_style(ws, 1, len(_BET_COLUMNS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_BET_COLUMNS))}1"

    # Data rows
    for r_idx, bet in enumerate(bets, start=2):
        tier = _str(bet.get("tier", "C"))
        decision = _str(bet.get("decision", ""))
        bg = _TIER_FILL.get(tier, _ALT_FILL)

        values = [
            _str(bet.get("match_id", "")),
            _str(bet.get("market", "")),
            _num(bet.get("probability")),
            _num(bet.get("odds")),
            _num(bet.get("ev")),
            _num(bet.get("kelly")),
            _num(bet.get("kelly_stake")),
            tier,
            _num(bet.get("confidence")),
            _num(bet.get("agreement")),
            _num(bet.get("market_edge")),
            decision,
            _num(bet.get("over_25")),
            _num(bet.get("btts_yes")),
        ]
        for col_idx, (val, (_, _, fmt, _)) in enumerate(
            zip(values, _BET_COLUMNS), start=1
        ):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.number_format = fmt
            cell.fill = _make_fill(bg)
            cell.alignment = Alignment(horizontal="center")

    _auto_width(ws, _BET_COLUMNS)


def _write_summary_sheet(ws, bets: list[dict]) -> None:
    """Fill the Summary worksheet with league breakdown."""
    from collections import defaultdict

    for col_idx, (header, _, _, _) in enumerate(_SUMMARY_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    _header_style(ws, 1, len(_SUMMARY_COLUMNS))

    acc: dict[str, dict] = defaultdict(
        lambda: {"bets": 0, "ev_sum": 0.0, "kelly_sum": 0.0, "bet_cnt": 0}
    )
    for b in bets:
        league = (
            _str(b.get("league"))
            or _str((b.get("extras") or {}).get("league"))
            or "Unknown"
        )
        acc[league]["bets"] += 1
        acc[league]["ev_sum"] += _num(b.get("ev"))
        acc[league]["kelly_sum"] += _num(b.get("kelly"))
        acc[league]["bet_cnt"] += 1 if b.get("decision") == "BET" else 0

    rows_data = []
    for league, a in sorted(acc.items()):
        n = a["bets"] or 1
        rows_data.append(
            {
                "league": league,
                "bets": a["bets"],
                "avg_ev": a["ev_sum"] / n,
                "avg_kelly": a["kelly_sum"] / n,
                "bet_rate": a["bet_cnt"] / n,
            }
        )

    alt = False
    for r_idx, row in enumerate(rows_data, start=2):
        bg = _ALT_FILL if alt else "FFFFFF"
        alt = not alt
        values = [
            _str(row["league"]),
            _int(row["bets"]),
            row["avg_ev"],
            row["avg_kelly"],
            row["bet_rate"],
        ]
        for col_idx, (val, (_, _, fmt, _)) in enumerate(
            zip(values, _SUMMARY_COLUMNS), start=1
        ):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.number_format = fmt
            cell.fill = _make_fill(bg)

    _auto_width(ws, _SUMMARY_COLUMNS)


def _write_metrics_sheet(ws, metrics: dict) -> None:
    """Fill the Metrics worksheet."""
    for col_idx, (header, _, _, _) in enumerate(_METRICS_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    _header_style(ws, 1, len(_METRICS_COLUMNS))

    kpis = [
        ("ROI", f"{_num(metrics.get('roi')):.2%}"),
        ("Yield", f"{_num(metrics.get('yield')):.2%}"),
        ("Hit Rate", f"{_num(metrics.get('hit_rate')):.2%}"),
        ("Total Bets", str(_int(metrics.get("total_bets")))),
        ("Total Profit", f"{_num(metrics.get('total_profit')):.2f}"),
        ("Total Staked", f"{_num(metrics.get('total_staked')):.2f}"),
        ("Max Drawdown", f"{_num(metrics.get('max_drawdown')):.2%}"),
        ("Brier Score", f"{_num(metrics.get('brier_score')):.4f}"),
        ("Log Loss", f"{_num(metrics.get('log_loss')):.4f}"),
    ]
    for r_idx, (kpi, val) in enumerate(kpis, start=2):
        ws.cell(row=r_idx, column=1, value=kpi).font = Font(bold=True)
        ws.cell(row=r_idx, column=2, value=val)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16


# ---------------------------------------------------------------------------
# Public class
# ---------------------------------------------------------------------------


class ExcelExporter:
    """
    Multi-sheet Excel exporter with full openpyxl formatting.

    Falls back to a basic pandas export when openpyxl is unavailable.
    """

    def __init__(self, output_dir: str = "outputs") -> None:
        self._dir = Path(output_dir)
        self._dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------

    def export(
        self,
        bets: list[dict],
        metrics: Optional[dict] = None,
        filename: Optional[str] = None,
        date_str: Optional[str] = None,
    ) -> str:
        """
        Export bets + metrics to a formatted Excel file.

        Returns
        -------
        str — absolute path of the written file.
        """
        today = date_str or date.today().isoformat()
        filename = filename or f"value_bets_{today}.xlsx"
        path = self._dir / filename
        metrics = metrics or {}

        if not _OPENPYXL_AVAILABLE:
            return self._fallback_export(bets, str(path))

        wb = openpyxl.Workbook()

        # --- Bets sheet ---
        ws_bets = wb.active
        ws_bets.title = "Bets"
        _write_bets_sheet(ws_bets, bets)

        # --- Summary sheet ---
        ws_sum = wb.create_sheet("Summary")
        _write_summary_sheet(ws_sum, bets)

        # --- Metrics sheet ---
        ws_met = wb.create_sheet("Metrics")
        _write_metrics_sheet(ws_met, metrics)

        # Workbook metadata
        wb.properties.title = "Football Quant Engine — Value Bets"
        wb.properties.subject = f"Report {today}"
        wb.properties.creator = "FootballQuantEngine v5"

        wb.save(str(path))
        return str(path.resolve())

    # ------------------------------------------------------------------
    # Legacy compatibility (used by DashboardView)
    # ------------------------------------------------------------------

    def export_value_bets(self, filepath: str, value_bets: list[dict]) -> None:
        """Write bets to an explicit filepath (backwards-compatible)."""
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)

        if not _OPENPYXL_AVAILABLE:
            self._fallback_export(value_bets, str(path))
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bets"
        _write_bets_sheet(ws, value_bets)
        wb.save(str(path))

    # ------------------------------------------------------------------
    # Fallback (pandas)
    # ------------------------------------------------------------------

    def _fallback_export(self, bets: list[dict], filepath: str) -> str:
        try:
            import pandas as pd

            df = pd.DataFrame(bets)
            df.to_excel(filepath, index=False)
        except Exception:
            pass
        return filepath
